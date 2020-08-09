
from PIL import Image
import cv2
import numpy as np
from tqdm import tqdm
import pytesseract
from pathlib import Path
import random
# from matplotlib import pyplot as plt
import pandas as pd
import random
import re
from openpyxl import load_workbook
import csv
import sys
import os
from scipy.ndimage import interpolation as inter
from fuzzywuzzy import fuzz
import shutil
from pdf2image import convert_from_path
import argparse

# construct the argument parse and parse the arguments
ap = argparse.ArgumentParser()
ap.add_argument("-i", "--input-file", required=True,
                help="pdf input to be extracted")
ap.add_argument("-o", "--output-file", type=str, default="",
                help="output path to final excel")
ap.add_argument("-t", "--template", type=str, default="",
                help="path to template excel")
args = vars(ap.parse_args())
pdf_input = args["input_file"]
try:
    excel_template = args["template"]
    1/len(excel_template)
except:
    excel_template = 'template.xlsx'
excel_output = args["output_file"]
output_xl = excel_output
print(excel_template)
#convert the pdf into a temporary jpeg file
class  ExtractImage:
  def __init__(self, filepath, outputdir):
    pages = convert_from_path(filepath, 400)
    outputfile = os.path.join(outputdir, 'output.jpg')
    pagenum = 0
    for page in pages:
      pagenum += 1
      page.save(outputfile, 'JPEG')
      break

temp_output_dir = './tmp'


#deskew the image
def deskew(image, delta=1, limit=5):
    def determine_score(arr, angle):
        data = inter.rotate(arr, angle, reshape=False, order=0)
        histogram = np.sum(data, axis=1)
        score = np.sum((histogram[1:] - histogram[:-1]) ** 2)
        return histogram, score

    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1] 

    scores = []
    angles = np.arange(-limit, limit + delta, delta)
    for angle in angles:
        histogram, score = determine_score(thresh, angle)
        scores.append(score)

    best_angle = angles[scores.index(max(scores))]

    (h, w) = image.shape[:2]
    center = (w // 2, h // 2)
    M = cv2.getRotationMatrix2D(center, best_angle, 1.0)
    rotated = cv2.warpAffine(image, M, (w, h), flags=cv2.INTER_CUBIC, \
              borderMode=cv2.BORDER_REPLICATE)

    return rotated




ExtractImage(pdf_input,temp_output_dir)
file = os.path.join(temp_output_dir,'output.jpg')



img_col = cv2.imread(file)



img_col = deskew(img_col)
img = cv2.cvtColor(img_col, cv2.COLOR_BGR2GRAY)

#Blur the image
img = cv2.GaussianBlur(img,(3,3),0)

#thresholding the image to a binary image
img_bin = cv2.adaptiveThreshold(img,255,cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY,11,2)

#inverting the image 
img_bin = 255-img_bin

# cv2.imwrite('/content/cv_inverted.png',img_bin)
# Plotting the image to see the output
# plotting = plt.imshow(img_bin,cmap='gray')
# plt.show()

# defining the kernel functions parameters for detecting horizontal and vertical lines
kernel_len = np.array(img).shape[1]//100
# Defining a vertical kernel to detect all vertical lines of image 
ver_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, kernel_len))
# Defining a horizontal kernel to detect all horizontal lines of image
hor_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (kernel_len, 1))
# A kernel of 2x2
kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))

#Use vertical kernel to detect and save the vertical lines in a jpg
image_1 = cv2.erode(img_bin, ver_kernel, iterations=3)
vertical_lines = cv2.dilate(image_1, ver_kernel, iterations=3)
# cv2.imwrite("/content/vertical.jpg",vertical_lines)
#Plot the generated image
# plotting = plt.imshow(image_1,cmap='gray')
# plt.show()

#Use horizontal kernel to detect and save the horizontal lines in a jpg
image_2 = cv2.erode(img_bin, hor_kernel, iterations=3)
horizontal_lines = cv2.dilate(image_2, hor_kernel, iterations=3)
# cv2.imwrite("/content/horizontal.jpg",horizontal_lines)
#Plot the generated image
# plotting = plt.imshow(image_2,cmap='gray')
# plt.show()

# Combine horizontal and vertical lines in a new third image, with both having same weight.
img_vh = cv2.addWeighted(vertical_lines, 0.5, horizontal_lines, 0.5, 0.0)
#Eroding and thesholding the image
img_vh = cv2.erode(~img_vh, kernel, iterations=2)
thresh, img_vh = cv2.threshold(img_vh,128,255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
# cv2.imwrite("img_vh.jpg", img_vh)
bitxor = cv2.bitwise_xor(img,img_vh)
bitnot = cv2.bitwise_not(bitxor)
#Plotting the generated image
# plotting = plt.imshow(bitnot,cmap='gray')
# plt.show()

# Detect contours for following box detection
contours, hierarchy = cv2.findContours(img_vh, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
def sort_contours(cnts, method="left-to-right"):
    # initialize the reverse flag and sort index
    reverse = False
    i = 0
    # handle if we need to sort in reverse
    if method == "right-to-left" or method == "bottom-to-top":
      reverse = True
    # handle if we are sorting against the y-coordinate rather than
    # the x-coordinate of the bounding box
    if method == "top-to-bottom" or method == "bottom-to-top":
      i = 1
    # construct the list of bounding boxes and sort them from top to
    # bottom
    boundingBoxes = [cv2.boundingRect(c) for c in cnts]
    (cnts, boundingBoxes) = zip(*sorted(zip(cnts, boundingBoxes),
    key=lambda b:b[1][i], reverse=reverse))
    # return the list of sorted contours and bounding boxes
    return (cnts, boundingBoxes)

# Sort all the contours by top to bottom.
contours, boundingBoxes = sort_contours(contours, method="top-to-bottom")

#Creating a list of heights for all detected boxes
heights = [boundingBoxes[i][3] for i in range(len(boundingBoxes))]
#Get mean of heights
mean = np.mean(heights)

#Create list box to store all boxes in  
box = []
# Get position (x,y), width and height for every contour and show the contour on image
for c in contours:
    x, y, w, h = cv2.boundingRect(c)
    if ((w<1000 and h<500) and (w/h < 20 and w/h > 0.05)):
        image = cv2.rectangle(img_col,(x,y),(x+w,y+h),(random.randint(0, 255),random.randint(0, 255),random.randint(0, 255)),4)
        box.append([x,y,w,h])
# cv2.imwrite("/content/image_cont.jpg", image)
# plotting = plt.imshow(image,cmap='gray')
# plt.show()

#Creating two lists to define row and column in which cell is located
try:
    row=[]
    column=[]
    j=0
    #Sorting the boxes to their respective row and column
    for i in range(len(box)):
        if (i==0):
            column.append(box[i])
            previous=box[i]
        else:
            # if (box[i][1]<=previous[1]+mean/2):
            if (box[i][1]<=previous[1]+mean/5):
                column.append(box[i])
                previous=box[i]
                if(i==len(box)-1):
                    row.append(column)
            else:
                row.append(column)
                column=[]
                previous = box[i]
                column.append(box[i])

    #calculating maximum number of cells
    countcol = 0
    max_col_row = 0
    for i in range(len(row)):
        currcount = len(row[i])
        if currcount > countcol:
            countcol = currcount
            max_col_row = i

    #Retrieving the center of each column
    # center = [int(row[i][j][0]+row[i][j][2]/2) for j in range(len(row[i])) if row[0]]
    center = [int(row[max_col_row][j][0]+row[max_col_row][j][2]/2) for j in range(len(row[max_col_row]))]
    center=np.array(center)
    center.sort()

    #Regarding the distance to the columns center, the boxes are arranged in respective order

    finalboxes = []
    for i in range(len(row)):
        lis=[]
        for k in range(countcol):
            lis.append([])
        for j in range(len(row[i])):
            diff = abs(center-(row[i][j][0]+row[i][j][2]/4))
            minimum = min(diff)
            indexing = list(diff).index(minimum)
            lis[indexing].append(row[i][j])
        finalboxes.append(lis)

    #from every single image-based cell/box the strings are extracted via pytesseract and stored in a list
    outer=[]
    for i in tqdm(range(len(finalboxes))):
        for j in range(len(finalboxes[i])):
            inner=''
            if(len(finalboxes[i][j])==0):
                outer.append(' ')
            else:
                for k in range(len(finalboxes[i][j])):
                    y,x,w,h = finalboxes[i][j][k][0],finalboxes[i][j][k][1], finalboxes[i][j][k][2],finalboxes[i][j][k][3]
                    finalimg = bitnot[x:x+h, y:y+w]
                    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 1))
                    border = cv2.copyMakeBorder(finalimg,2,2,2,2,   cv2.BORDER_CONSTANT,value=[255,255])
                    resizing = cv2.resize(border, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
                    dilation = cv2.dilate(resizing, kernel,iterations=1)
                    erosion = cv2.erode(dilation, kernel,iterations=1)

                    out = pytesseract.image_to_string(erosion)
                    if(len(out)==0):
                        out = pytesseract.image_to_string(erosion, config='--psm 3')
                    inner = inner +" "+ out
                outer.append(inner)

    #Creating a dataframe of the generated OCR list
    arr = np.array(outer)
    dataframe = pd.DataFrame(arr.reshape(len(row),countcol))
    # print(dataframe)
    data = dataframe.style.set_properties(align="left")

    # Assume first row is header row
    headers = list(dataframe.iloc[0])
    res = []
    for sub in headers: 
        res.append(sub.replace("\n", " ")) 

    dataframe.columns = res
    dataframe.drop(0, inplace=True)

    # dataframe.to_csv("output.csv")

    template = pd.read_excel(excel_template)

    template = template.iloc[15:16, :]

    template.index = np.arange(len(template))
    template.columns = template.iloc[0, :]
    template.drop(0, inplace=True)

    cols = []

    # S.No
    names = ["S.No", "Serial No.", "S.Num", "Serial Number"]
    cols.append(names)

    # Product_ID
    names = ["Product ID", "Prod ID", "Prod Code", "Item Code", "Item ID", "ID"]
    cols.append(names)

    # SKU
    names = ["SKU"]
    cols.append(names)
    # HSN
    names = ["HSN", "HS Code", "HS"]
    cols.append(names)

    # Title
    names = ["Title", "Description", "Desc", "Name", "Product Name"]
    cols.append(names)

    # Quantity
    names = ["Quantity", "Qty"]
    cols.append(names)

    # Unit Price
    names = ["Unit Price", "Unit", "Price per unit"]
    cols.append(names)

    # Excise
    names = ["Excise Duty", "Excise", "Duty"]
    cols.append(names)

    # Discount %
    names = ["Discount percent", "Discount", "Disc"]
    cols.append(names)

    # SGST
    names = ["SGST"]
    cols.append(names)

    # CGST
    names = ["CGST"]
    cols.append(names)

    # IGST
    names = ["IGST"]
    cols.append(names)

    # Cess Percent
    names = ["Cess", "Cess %", "Cess percent"]
    cols.append(names)

    # TCS Percent
    names = ["TCS", "TCS %", "TCS Percentage"]
    cols.append(names)

    # Total Amount
    names = ["Total", "Total Amount", "Invoice Value"]
    cols.append(names)

    # App %
    names = ["App %"]
    cols.append(names)

    def map_to_template(inword, searchlist):
      for i, list in enumerate(searchlist):
        for word in list:
          word = word.lower()
          inword = inword.lower()
          # if (inword.lower() in word.lower()) or (word.lower() in inword.lower()):
          if fuzz.token_set_ratio(word, inword) > 75:
            return i

      return None


    data_headers = dataframe.columns
    mapped_cols = set()
    for head in data_headers:
      if head == '' or head == ' ':
        continue
      col_num = map_to_template(head, cols)
      if col_num is not None and col_num not in mapped_cols:
        dataframe.rename(columns = {head : template.columns[col_num]}, inplace=True)
        mapped_cols.add(col_num)

    input_cols = set(dataframe.columns)
    target_cols = set(template.columns)
    common_cols = list(input_cols.intersection(target_cols))

    template[common_cols] = dataframe[common_cols].values

    template

    # template.to_excel("/content/template.xlsx", index=False)
    #function that appends onto existing excel file
    def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,startcol=0,
                           truncate_sheet=False, 
                           **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        Parameters:
          filename : File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
          df : dataframe to save to workbook
          sheet_name : Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
          startrow : upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
          truncate_sheet : truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
          to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]

        Returns: None
        """
        from openpyxl import load_workbook

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl')

        # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
        try:
            FileNotFoundError
        except NameError:
            FileNotFoundError = IOError


        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow,startcol=startcol ,**to_excel_kwargs)

        # save the workbook
        writer.save()

    # Template excel sheet
    target_template_xl = excel_template

    template.drop('S.No',axis=1,inplace=True)
    output_xl = excel_output

    shutil.copy(target_template_xl, output_xl)

    append_df_to_excel(output_xl, template, startrow=16, startcol=1, index=False)

except Exception as e:
    print('notable')
    shutil.copy(excel_template,excel_output)


print('Done with table')


#gets the edit distance (minimum number of character insertions, deletions, additions) for two strings
def editDist(str1, str2): 
    # Create a table to store results of subproblems 
    m = len(str1)
    n = len(str2)
    dp = [[0 for x in range(n + 1)] for x in range(m + 1)]   
    for i in range(m + 1): 
        for j in range(n + 1): 
            if i == 0: 
                dp[i][j] = j    # Min. operations = j 

            elif j == 0: 
                dp[i][j] = i    # Min. operations = i 

            elif str1[i-1] == str2[j-1]: 
                dp[i][j] = dp[i-1][j-1] 
            else: 
                dp[i][j] = 1 + min(dp[i][j-1],        # Insert 
                                   dp[i-1][j],        # Remove 
                                   dp[i-1][j-1])    # Replace 
  
    return dp[m][n] 




def isConnected2(x1,delx1, x2, delx2,leeway=0): ##used for vertical connectivity
    mid1 = x1 +delx1/2
    mid2 = x2 +delx2/2
    return (mid1>=x2 and mid1<=(x2+delx2) or mid2>=x1 and mid2<=(x1+delx1))




def isConnected(x1,delx1, x2, delx2,leeway=0):
    if(x1<x2):
        return x2 <= x1+delx1 +leeway
    return x1 <= x2 + delx2 + leeway




def get_actual_doc(arr):  #crops out the central document
    xl = arr.shape[1]
    yt = arr.shape[0]
    yb=0
    xr=0
    imgray = cv2.cvtColor(arr,cv2.COLOR_BGR2GRAY)
    ret,thresh = cv2.threshold(imgray,230,255,cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    contours,_ = cv2.findContours(thresh,cv2.RETR_LIST,cv2.CHAIN_APPROX_SIMPLE)
    for i in range(len(contours)):
        cnt = contours[i]
        x,y,w,h = cv2.boundingRect(cnt)
        if(w*h >0.8*arr.shape[0]*arr.shape[1] or w*h<=arr.shape[0]*arr.shape[1]*0.0001):
            continue
        xr = max(xr,x+w)
        xl = min(xl,x)
        yt = min(yt,y)
        yb = max(yb,y+h)
    return arr[yt:yb,xl:xr]


# def get_properly_oriented_image(image, delta=1, limit=5):
#     def determine_score(arr, angle):
#         data = inter.rotate(arr, angle, reshape=False, order=0)
#         histogram = np.sum(data, axis=1)
#         score = np.sum((histogram[1:] - histogram[:-1]) ** 2)
#         return histogram, score

#     gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
#     thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1] 

#     scores = []
#     angles = np.arange(-limit, limit + delta, delta)
#     for angle in angles:
#         histogram, score = determine_score(thresh, angle)
#         scores.append(score)

#     best_angle = angles[scores.index(max(scores))]

#     (h, w) = image.shape[:2]
#     center = (w // 2, h // 2)
#     M = cv2.getRotationMatrix2D(center, best_angle, 1.0)
#     rotated = cv2.warpAffine(image, M, (w, h), flags=cv2.INTER_CUBIC,               borderMode=cv2.BORDER_REPLICATE)

#     return rotated




def get_table(img):   #extracts the table from the image
    if(type(img) == str):
        img = cv2.imread(img)
    if(len(img.shape)>2):
        imgray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
    else:
        imgray = img
#     _,thresh = cv2.threshold(imgray,230,255,cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    imgray = cv2.GaussianBlur(imgray,(3,3),0)
    thresh = cv2.adaptiveThreshold(imgray,255,cv2.ADAPTIVE_THRESH_GAUSSIAN_C,cv2.THRESH_BINARY,101,2)
    thresh = 255-thresh
    # Length(width) of kernel as 100th of total width
    kernel_len = np.array(img).shape[1]//80
    kernel_len0 = np.array(img).shape[0]//40 ##was originally 40
    kernel_len1 = np.array(img).shape[1]//25 
    
    
    # Defining a vertical kernel to detect all vertical lines of image 
    ver_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, kernel_len0))
    
    # Defining a horizontal kernel to detect all horizontal lines of image
    hor_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (kernel_len1, 1))
    
    # A kernel of 2x2
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    
    image_1 = cv2.erode(thresh, ver_kernel, iterations=3)
    vertical_lines = cv2.dilate(image_1, ver_kernel, iterations=3)
    
    image_2 = cv2.erode(thresh, hor_kernel, iterations=3)
    horizontal_lines = cv2.dilate(image_2, hor_kernel, iterations=3)
    
    img_vh = cv2.addWeighted(vertical_lines, 0.5, horizontal_lines, 0.5, 0.0)
    #Eroding and thesholding the image
    img_vh = cv2.erode(~img_vh, kernel, iterations=1)
    thresh, img_vh = cv2.threshold(img_vh,128,255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    return img_vh




def show_text(fname):
    if(type(fname) == str):
        img = Image.open(fname)
    else:
        img = fname
    arr2 = np.array(img)
    text_boxes = []
    arr2 = get_actual_doc(arr2)
    arr = arr2.copy()
    arr2[get_table(arr2) == 0] = np.percentile(arr2,80)
    scl = 2
    arr2 = cv2.resize( arr2, (int(arr2.shape[1]*scl),int(arr2.shape[0]*scl)), interpolation = cv2.INTER_LANCZOS4)
    imgray = cv2.cvtColor(arr2,cv2.COLOR_BGR2GRAY)
    ret,thresh = cv2.threshold(imgray,230,255,cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    contours,_ = cv2.findContours(thresh,cv2.RETR_LIST,cv2.CHAIN_APPROX_SIMPLE)
    for i in range(len(contours)):
        cnt = contours[i]
        x,y,w,h = cv2.boundingRect(cnt)
        if(not(w*h >0.9*arr2.shape[0]*arr2.shape[1] or (h/w>4 and h>arr2.shape[0]/100))):
            text_boxes.append((x,y,w,h))

    adj_list = [[] for i in range(len(text_boxes))]
    visited = np.zeros(len(text_boxes))
    
    for i in range(len(text_boxes)):    ##building the adjacency matrix
        for j in range(i+1,len(text_boxes)):
            x1,y1,dx1,dy1 = text_boxes[i]
            x2,y2,dx2,dy2 = text_boxes[j]
            if (isConnected(x1,dx1,x2,dx2,4*scl) and isConnected2(y1,dy1,y2,dy2)): #check if two boxes are connected
                adj_list[i].append(j)
                adj_list[j].append(i)

    final_text_boxes=[]
    def dfs(v):
        if(visited[v]):
            return ()
        visited[v]=1
        x,y,w,h = text_boxes[v] 
        for i in adj_list[v]:
            if(visited[i]):
                continue
            x2,y2,w2,h2 = dfs(i)
            xtmp = x
            ytmp = y
            wtmp = w
            htmp = h
            xtmp = min(x,x2)
            ytmp = min(y,y2)
            wtmp = max(x+w,x2+w2)
            wtmp = wtmp - xtmp
            htmp = max(y+h,y2+h2)
            htmp = htmp - ytmp
            x=xtmp
            y=ytmp
            h=htmp
            w=wtmp
        return (x,y,w,h)
    for i in (range(len(text_boxes))):
        res = dfs(i)
        if(len(res)):
            final_text_boxes.append(res)

    words = []
#     words.append(['x','y','w','h','word/block'])
    i=0
    txt_image = np.zeros_like(arr2)
    arr = arr2.copy()
    dx = 0*scl
    dy = 0*scl
    return pytesseract.image_to_data(arr,lang='eng',output_type=pytesseract.Output.DATAFRAME)
    




df = show_text(img_col)
df.dropna(inplace =True)
df2 = df.copy()
df.drop(['level','page_num','block_num','par_num','line_num','word_num','conf'],axis = 1,inplace = True)
df.columns = ['x','y','w','h','word']
df['word'] = df['word'].apply(lambda x: x.split(':'))
df = df.explode('word')
df.index = np.arange(len(df))
df2 = df.copy()
final_key_value_pairs = {}


def get_val(df, key_name,regex_string): 
  key_df = df
  key_df['dist'] = df['word'].apply(lambda x: editDist(x,key_name))
  key_df.sort_values('dist',inplace = True)
  key_df.index = np.arange(len(key_df))
  x_key,y_key,w_key,h_key,word,_=list(df.iloc[0,:])
  key_dist_string = key_name +'_Distance'
  df[key_dist_string] = df.apply(lambda x: (x['x'] - x_key)**2 + (y_key - x['y'])**2,axis=1)
  key_df = df.sort_values(key_dist_string)
  key_df.index = np.arange(len(key_df))
  key_df.word = key_df.word.apply(lambda x: x.replace("|",""))
  return key_df[(key_df.word.apply(lambda x: re.search(regex_string, x)!=None))]


def get_val_GSTIN(df, key_name,regex_string): 
  key_df = df
  
  key_df.index = np.arange(len(key_df))
  key_df.word = key_df.word.apply(lambda x: x.split('/'))
  key_df = key_df.explode('word')
  key_df.index = np.arange(len(key_df))
  df = key_df
  key_df['dist'] = df['word'].apply(lambda x: editDist(x,key_name))
  key_df.sort_values('dist',inplace = True)
  final_df = None
  key_df2 = key_df.copy()
  df2 = df.copy()
  for i in range(len(key_df)):
    key_df = key_df2.copy()
    df = df2.copy()
    x_key,y_key,w_key,h_key,word,dist=list(df.iloc[i,:])
    if(dist >2):
      break
    key_dist_string = key_name +'_Distance'
    df[key_dist_string] = df.apply(lambda x: (x['x'] - x_key)**2 + (y_key - x['y'])**2,axis=1)
    key_df = df.sort_values(key_dist_string)
    key_df.index = np.arange(len(key_df))
    key_df.word = key_df.word.apply(lambda x: x.replace("|",""))
    matching_row = key_df[(key_df.word.apply(lambda x: re.search(regex_string, x)!=None))].iloc[0,:]
    if type(final_df) == None:
      final_df = matching_row
    else:
      final_df = pd.concat([final_df,matching_row],axis=1)
  final_df = final_df.T.sort_values('y')
  gstin_vals = {}
  final_df.index = np.arange(len(final_df))
  if(len(final_df) == 1):
    gstin_vals['Seller_GSTIN'] = final_df.iloc[0,4]
  else:
    gstin_vals['Seller_GSTIN'] = final_df.iloc[0,4]
    gstin_vals['Buyer_GSTIN'] = final_df.iloc[len(final_df)-1,4]
  return gstin_vals


df = df2.copy()
gstin_vals = get_val_GSTIN(df,'GSTIN', '^(?=.*[0-9])(?=.*[a-zA-Z])(?=\S+$).{10,18}$')
final_key_value_pairs.update(gstin_vals)
# 'GSTIN', '^(?=.*[0-9])(?=.*[a-zA-Z])(?=\S+$).{10,18}$'
# 'Date',  '^(0[1-9]|[12][0-9]|3[01])[- /.](0[1-9]|1[012])[- /.](19|20)\d\d$'

def get_val_State(df, key_name,state_names): 
  
  key_df = df
  state_names_lower = [s.lower() for s in state_names]
  key_df.index = np.arange(len(key_df))
  key_df.word = key_df.word.apply(lambda x: x.split('/'))
  key_df = key_df.explode('word')
  key_df.index = np.arange(len(key_df))
  df = key_df
  key_df['dist'] = df['word'].apply(lambda x: editDist(x,key_name))
  key_df.sort_values('dist',inplace = True)
  final_df = None
  key_df2 = key_df.copy()
  df2 = df.copy()
  for i in range(len(key_df)):
    key_df = key_df2.copy()
    df = df2.copy()
    x_key,y_key,w_key,h_key,word,dist=list(df.iloc[i,:])
    if(dist >2):
      break
    key_dist_string = key_name +'_Distance'
    df[key_dist_string] = df.apply(lambda x: (x['x'] - x_key)**2 + (y_key - x['y'])**2,axis=1)
    key_df = df.sort_values(key_dist_string)
    key_df.index = np.arange(len(key_df))
    key_df.word = key_df.word.apply(lambda x: x.replace("|",""))
#     matching_row = key_df[(key_df.word.apply(lambda x: re.search(regex_string, x)!=None))].iloc[0,:]
    try:
        matching_row = key_df[key_df.word.apply(lambda x: ''.join(e.lower() for e in x if e.isalpha()) in state_names_lower)].iloc[0,:]
    except:
        matching_row = key_df.iloc[0,:]
        matching_row['word'] = 'NONE'
    if type(final_df) == None:
      final_df = matching_row
    else:
      final_df = pd.concat([final_df,matching_row],axis=1)
  final_df = final_df.T.sort_values('y')
  state_vals = {}
  final_df.index = np.arange(len(final_df))
  state_vals['Seller State'] = final_df.iloc[0,4]
  return state_vals





state_names = ["Andhra","Arunachal","Assam","Bihar","Chhattisgarh","Goa","Gujarat","Haryana","Himachal","Jammu and Kashmir","Jharkhand","Karnataka","Kerala","Madhya ","Maharashtra","Manipur",
                "Meghalaya","Mizoram","Nagaland","Odisha","Punjab","Rajasthan","Sikkim","Tamil Nadu","Telangana","Tripura","Uttarakhand","Uttar","West Bengal","Andaman and Nicobar Islands",
                "Chandigarh","Dadra and Nagar Haveli","Daman and Diu","Delhi","Lakshadweep","Puducherry"]


df = df2.copy()
state_val = get_val_State(df,'State',state_names)
final_key_value_pairs.update(state_val)

def get_val_Date(df, key_name,regex_string): 
  key_df = df
  key_df.index = np.arange(len(key_df))
  dates_df = key_df[(key_df.word.apply(lambda x: re.search(regex_string, x)!=None))]
  final_df = None
  key_df2 = key_df.copy()
  df2 = df.copy()
  for i in range(len(dates_df)):
    key_df = key_df2.copy()
    df = df2.copy()
    x_key,y_key,w_key,h_key,word=list(dates_df.iloc[i,:])
    key_df['Distance'] = key_df.apply(lambda x: (x['x'] - x_key)**2 + (y_key - x['y'])**2,axis=1)
    key_df = key_df.sort_values('Distance')
    key_df.index = np.arange(len(key_df))
    key_df.word = key_df.word.apply(lambda x: x.replace("|",""))
    matching_row = key_df.iloc[0,:]
    if type(final_df) == None:
      final_df = matching_row
    else:
      final_df = pd.concat([final_df,matching_row],axis=1)
  final_df = final_df.T.sort_values('y')
  date_vals = {}
  final_df.index = np.arange(len(final_df))
  date_vals['Invoice Date'] = final_df.iloc[0,4]
  return date_vals




df = df2.copy()
# get_val_Date(df,'Date','^(0[1-9]|[12][0-9]|3[01])[- /.](0[1-9]|1[012])[- /.](19|20)\d\d$')
date_vals = get_val_Date(df,'Date',"(0?[1-9]|[12]\d|30|31)[^\w\d\r\n:](0?[1-9]|1[0-2])[^\w\d\r\n:](\d{4}|\d{2})")
final_key_value_pairs.update(date_vals)




df = df2.copy()
d=get_val(df,'','^PO[0-9]{5,10}$')
try:
    po_vals = {'PO Number':d.iloc[0,4]}
except:
    df = df2.copy()
    d= get_val(df,'PO','^[0-9]{5,10}$')
    try: 
        po_vals = {'PO Number':d.iloc[0,4]}
    except:
        po_vals = {'PO Number': "NONE"}
final_key_value_pairs.update(po_vals)




def get_val_invoice_no(df, key_name,regex_string): 
  key_df = df
  df_word_copy = df.word.copy()
  df_word_next = df.word[1:]
  df_word_next.index = np.arange(len(df_word_next))
  df.word = df.word[:-1] +' '+df_word_next
  key_df = df
  key_df['dist'] = df['word'].apply(lambda x: editDist(str(x),key_name))
  key_df.sort_values('dist',inplace = True)
  key_df.index = np.arange(len(key_df))
  x_key,y_key,w_key,h_key,word,_=list(df.iloc[0,:])
  df.word = df.word.apply(lambda x:str(x).split(' ')[0])
  key_dist_string = key_name +'_Distance'
  df[key_dist_string] = df.apply(lambda x: (x['x'] - x_key)**2 + (y_key - x['y'])**2,axis=1)
  key_df = df.sort_values(key_dist_string)
  key_df.index = np.arange(len(key_df))
  key_df.word = key_df.word.apply(lambda x: str(x).replace("|",""))
  return key_df[(key_df.word.apply(lambda x: re.search(regex_string, str(x))!=None))]




df = df2.copy()
d = get_val_invoice_no(df,'invoice no','^[a-zA-Z]?[0-9]{4,10}$')
try:
    invoice_no_vals = {'Invoice Number':d.iloc[0,4]}
except:
    invoice_no_vals = {'Invoice Number': "NONE"}
final_key_value_pairs.update(invoice_no_vals)
print(final_key_value_pairs)




cur = Path.cwd()
workbook = load_workbook(filename=excel_output)
workbook.sheetnames
spreadsheet = workbook.active
spreadsheet['E3'] = final_key_value_pairs['Seller State']
spreadsheet['E11'] = final_key_value_pairs['Seller_GSTIN']
spreadsheet['M13'] = final_key_value_pairs['Buyer_GSTIN']
spreadsheet['M10'] = final_key_value_pairs['PO Number']
spreadsheet['M3'] = final_key_value_pairs['Invoice Number']
spreadsheet['M4'] = final_key_value_pairs['Invoice Date']
workbook.save(filename=excel_output)





